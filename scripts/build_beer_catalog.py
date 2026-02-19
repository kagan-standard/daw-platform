#!/usr/bin/env python3
"""
Beer Catalog Pipeline — Phase 3.1
Reads local data files, deduplicates, outputs CSVs for breweries, beers, beer_styles, flavor_descriptors.
No external API calls. Requires data/*.xlsx and data/*.csv in repo data/ directory.
"""
import csv
import re
import sys
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl

# Paths
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DATA_DIR = REPO_ROOT / "data"
OUT_DIR = DATA_DIR / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------------------------------------------------------------
# 2B: Utility functions
# -----------------------------------------------------------------------------

def slugify(name: str) -> str:
    """URL-safe slug: 'Sierra Nevada Brewing Co.' -> 'sierra-nevada-brewing-co'"""
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', name.lower())).strip('-')

def normalize_name(name: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace. For dedup matching."""
    if not name:
        return ""
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9\s]', '', name.lower())).strip()

def normalize_brewery(name: str) -> str:
    """Like normalize_name but also strips common suffixes."""
    if not name:
        return ""
    name = re.sub(r'\s+', ' ', name.strip().lower())
    for suffix in [
        'brewing company', 'brewing co.', 'brewing co', 'brewery',
        'beer company', 'beer co.', 'beer co', 'brew co',
        'breweries', 'brauerei', 'brasserie', 'brouwerij',
        'llc', 'inc.', 'inc', 'ltd.', 'ltd', 'co.', 'gmbh',
        'obergärige hausbrauerei gmbh'
    ]:
        name = re.sub(rf'\s*{re.escape(suffix)}\s*$', '', name, flags=re.IGNORECASE)
    return re.sub(r'[^a-z0-9\s]', '', name).strip()

def fix_mojibake(text: str) -> str:
    """Fix common UTF-8-as-Latin-1 encoding artifacts."""
    if not text:
        return text
    replacements = {
        'Ã©': 'é', 'Ã¨': 'è', 'Ã¶': 'ö', 'Ã¼': 'ü', 'Ã±': 'ñ',
        'Ã¤': 'ä', 'Ã³': 'ó', 'Ã¡': 'á', 'Ã­': 'í', 'Ã§': 'ç',
        'Ã¢': 'â', 'Ãª': 'ê', 'Ã®': 'î', 'Ã´': 'ô', 'Ã»': 'û',
        'Ã¨re': 'ère', 'BiÃ¨re': 'Bière',
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text

def safe_float(v):
    try:
        return round(float(v), 2) if v is not None and str(v).strip() != '' else None
    except Exception:
        return None

def safe_int(v):
    try:
        return int(float(v)) if v is not None and str(v).strip() != '' else None
    except Exception:
        return None

def csv_val(v):
    """Convert Python value to CSV-safe string. None -> empty string (not 'None')."""
    if v is None:
        return ''
    if isinstance(v, float):
        return str(v)
    return str(v)


# -----------------------------------------------------------------------------
# 2H: Style mapping (BJCP-derived) + 2I: BJCP styles seed data
# -----------------------------------------------------------------------------

# BJCP-derived style list (~87 styles): name, category, description, abv_min, abv_max, ibu_min, ibu_max
BJCP_STYLES = [
    ("American Light Lager", "Lager", "Light American lager", 2.8, 4.2, 8, 12),
    ("American Lager", "Lager", "American standard lager", 4.2, 5.3, 8, 18),
    ("Cream Ale", "Hybrid", "Cream ale", 4.2, 5.6, 8, 20),
    ("American Wheat Beer", "Wheat", "American wheat", 4.0, 5.5, 15, 30),
    ("International Pale Lager", "Lager", "International pale lager", 4.5, 6.0, 18, 25),
    ("International Amber Lager", "Lager", "International amber lager", 4.6, 6.0, 8, 25),
    ("International Dark Lager", "Lager", "International dark lager", 4.2, 6.0, 8, 20),
    ("Czech Pale Lager", "Lager", "Czech pale lager", 3.0, 4.1, 20, 35),
    ("Czech Premium Pale Lager", "Lager", "Czech premium pale lager", 4.2, 5.8, 30, 45),
    ("Czech Amber Lager", "Lager", "Czech amber lager", 4.4, 5.8, 20, 35),
    ("Czech Dark Lager", "Lager", "Czech dark lager", 4.4, 5.8, 18, 34),
    ("Munich Helles", "Lager", "Munich Helles", 4.7, 5.4, 16, 22),
    ("Festbier", "Lager", "Festbier", 5.8, 6.3, 18, 25),
    ("Helles Bock", "Bock", "Helles Bock", 6.3, 7.4, 23, 35),
    ("German Leichtbier", "Lager", "German Leichtbier", 2.4, 3.6, 15, 28),
    ("Kolsch", "Hybrid", "Kolsch", 4.4, 5.2, 18, 30),
    ("German Helles Export", "Lager", "German Helles Export", 5.0, 6.0, 20, 30),
    ("German Pils", "Lager", "German Pils", 4.4, 5.2, 22, 40),
    ("Marzen", "Lager", "Marzen", 5.8, 6.3, 18, 24),
    ("Rauchbier", "Lager", "Rauchbier", 5.1, 6.0, 20, 30),
    ("Dunkles Bock", "Bock", "Dunkles Bock", 6.3, 7.2, 20, 27),
    ("Eisbock", "Bock", "Eisbock", 9.0, 14.0, 25, 35),
    ("Baltic Porter", "Porter", "Baltic Porter", 6.5, 9.5, 20, 40),
    ("American Pale Ale", "Pale Ale", "American Pale Ale", 4.5, 5.4, 30, 50),
    ("American IPA", "IPA", "American IPA", 5.5, 7.5, 40, 70),
    ("Imperial IPA", "IPA", "Imperial IPA", 7.5, 10.0, 65, 100),
    ("Hazy IPA", "IPA", "Hazy IPA", 5.5, 7.5, 35, 70),
    ("British Golden Ale", "Pale Ale", "British Golden Ale", 3.8, 5.0, 20, 45),
    ("Australian Sparkling Ale", "Pale Ale", "Australian Sparkling Ale", 4.5, 6.0, 20, 35),
    ("English Pale Ale", "Pale Ale", "English Pale Ale", 4.0, 5.5, 20, 40),
    ("American Amber Ale", "Amber/Red", "American Amber Ale", 4.5, 6.2, 25, 45),
    ("California Common", "Amber/Red", "California Common", 4.5, 5.5, 30, 45),
    ("American Brown Ale", "Brown Ale", "American Brown Ale", 4.3, 6.2, 25, 45),
    ("English Brown Ale", "Brown Ale", "English Brown Ale", 4.0, 5.4, 20, 30),
    ("Brown Porter", "Porter", "Brown Porter", 4.0, 5.4, 18, 35),
    ("Robust Porter", "Porter", "Robust Porter", 4.8, 6.5, 25, 50),
    ("Irish Stout", "Stout", "Irish Stout", 4.0, 4.5, 25, 45),
    ("Sweet Stout", "Stout", "Sweet Stout", 4.0, 6.0, 20, 40),
    ("Oatmeal Stout", "Stout", "Oatmeal Stout", 4.2, 5.9, 25, 40),
    ("American Stout", "Stout", "American Stout", 5.0, 7.0, 35, 75),
    ("Imperial Stout", "Stout", "Imperial Stout", 8.0, 12.0, 50, 90),
    ("Irish Extra Stout", "Stout", "Irish Extra Stout", 5.5, 6.5, 35, 50),
    ("Milk Stout", "Stout", "Milk Stout", 4.0, 6.0, 15, 40),
    ("Scottish Light", "Amber/Red", "Scottish Light", 2.5, 3.2, 10, 20),
    ("Scottish Heavy", "Amber/Red", "Scottish Heavy", 3.2, 3.9, 12, 20),
    ("Scottish Export", "Amber/Red", "Scottish Export", 3.9, 5.0, 15, 30),
    ("Irish Red Ale", "Amber/Red", "Irish Red Ale", 4.0, 5.0, 18, 28),
    ("American Strong Ale", "Strong Ale", "American Strong Ale", 6.0, 10.0, 50, 100),
    ("English Barleywine", "Barleywine", "English Barleywine", 8.0, 12.0, 35, 70),
    ("American Barleywine", "Barleywine", "American Barleywine", 8.0, 12.0, 50, 100),
    ("Witbier", "Belgian", "Witbier", 4.5, 5.5, 8, 20),
    ("Belgian Pale Ale", "Belgian", "Belgian Pale Ale", 4.8, 5.5, 20, 30),
    ("Saison", "Belgian", "Saison", 5.0, 7.0, 20, 35),
    ("Belgian Blond Ale", "Belgian", "Belgian Blond Ale", 6.0, 7.5, 15, 30),
    ("Belgian Dubbel", "Belgian", "Belgian Dubbel", 6.0, 7.6, 15, 25),
    ("Belgian Tripel", "Belgian", "Belgian Tripel", 7.5, 9.5, 20, 40),
    ("Belgian Golden Strong", "Belgian", "Belgian Golden Strong", 7.5, 10.5, 22, 35),
    ("Belgian Dark Strong", "Belgian", "Belgian Dark Strong", 8.0, 12.0, 20, 35),
    ("Flanders Red", "Sour/Wild", "Flanders Red", 4.6, 6.5, 10, 25),
    ("Flanders Brown", "Sour/Wild", "Flanders Brown", 4.0, 8.0, 10, 25),
    ("Gueuze", "Sour/Wild", "Gueuze", 5.0, 8.0, 0, 15),
    ("Lambic", "Sour/Wild", "Lambic", 5.0, 6.5, 0, 10),
    ("Berliner Weisse", "Sour/Wild", "Berliner Weisse", 2.8, 3.8, 3, 8),
    ("Gose", "Sour/Wild", "Gose", 4.2, 4.8, 5, 12),
    ("American Wild Ale", "Sour/Wild", "American Wild Ale", 4.0, 8.0, 0, 30),
    ("Weissbier", "Wheat", "Weissbier", 4.3, 5.6, 8, 15),
    ("Dunkles Weissbier", "Wheat", "Dunkles Weissbier", 4.3, 5.6, 10, 18),
    ("Weizenbock", "Wheat", "Weizenbock", 6.5, 9.0, 15, 30),
    ("Ordinary Bitter", "Pale Ale", "Ordinary Bitter", 3.2, 3.8, 25, 35),
    ("Best Bitter", "Pale Ale", "Best Bitter", 3.8, 4.6, 25, 40),
    ("Strong Bitter", "Pale Ale", "Strong Bitter", 4.6, 6.2, 30, 50),
    ("English Mild", "Brown Ale", "English Mild", 3.0, 3.8, 10, 25),
    ("Old Ale", "Strong Ale", "Old Ale", 5.5, 9.0, 30, 55),
    ("English IPA", "IPA", "English IPA", 5.0, 7.5, 40, 60),
    ("Dark Mild", "Brown Ale", "Dark Mild", 3.0, 3.8, 10, 25),
    ("Doppelbock", "Bock", "Doppelbock", 7.0, 10.0, 16, 26),
    ("Maibock", "Bock", "Maibock", 6.3, 7.4, 23, 35),
    ("Schwarzbier", "Lager", "Schwarzbier", 4.4, 5.4, 22, 32),
    ("Vienna Lager", "Lager", "Vienna Lager", 4.7, 5.5, 18, 30),
    ("Altbier", "Amber/Red", "Altbier", 4.3, 5.5, 25, 50),
    ("Dusseldorf Altbier", "Amber/Red", "Dusseldorf Altbier", 4.5, 5.2, 35, 50),
    ("Kellerbier", "Lager", "Kellerbier", 4.7, 5.4, 20, 35),
    ("Kentucky Common", "Hybrid", "Kentucky Common", 4.0, 5.5, 15, 30),
    ("Pre-Prohibition Lager", "Lager", "Pre-Prohibition Lager", 4.5, 6.0, 25, 40),
    ("Pre-Prohibition Porter", "Porter", "Pre-Prohibition Porter", 4.5, 6.0, 20, 30),
    ("Rye Beer", "Hybrid", "Rye Beer", 4.0, 6.0, 25, 45),
    ("Fruit Beer", "Specialty", "Fruit Beer", 2.5, 7.0, 5, 70),
    ("Spice Beer", "Specialty", "Spice Beer", 2.5, 12.0, 5, 70),
    ("Smoke Beer", "Specialty", "Smoke Beer", 4.0, 6.0, 20, 40),
    ("Winter Warmer", "Strong Ale", "Winter Warmer", 5.5, 9.0, 20, 45),
    ("Honey Beer", "Specialty", "Honey Beer", 3.5, 7.0, 10, 40),
    ("Roggenbier", "Wheat", "Roggenbier", 4.5, 6.0, 15, 30),
    ("New Zealand Pilsner", "Lager", "New Zealand Pilsner", 4.5, 5.5, 25, 45),
    ("Belgian Single", "Belgian", "Belgian Single", 4.5, 5.5, 20, 35),
    ("Dark Lager", "Lager", "Dark Lager", 4.0, 6.0, 14, 28),
    ("Pale Lager", "Lager", "Pale Lager", 4.0, 6.0, 18, 35),
    ("Pilsner", "Lager", "Pilsner", 4.2, 5.5, 22, 40),
    ("Bock", "Bock", "Bock", 6.0, 7.5, 20, 30),
    ("Chocolate Stout", "Stout", "Chocolate Stout", 4.0, 6.0, 20, 40),
    ("Coffee Stout", "Stout", "Coffee Stout", 4.0, 7.0, 25, 50),
    ("Pale Ale", "Pale Ale", "Pale Ale", 4.0, 5.5, 20, 45),
    ("IPA", "IPA", "IPA", 5.0, 7.5, 40, 70),
    ("Porter", "Porter", "Porter", 4.0, 6.5, 18, 50),
    ("Stout", "Stout", "Stout", 4.0, 7.0, 25, 60),
    ("Barleywine", "Barleywine", "Barleywine", 8.0, 12.0, 35, 100),
]


def map_style_to_bjcp(style_raw: str | None) -> str | None:
    """Map raw style string to BJCP-derived style name (matches beer_styles.name)."""
    if not style_raw or not str(style_raw).strip():
        return None
    s = fix_mojibake(str(style_raw).strip().lower())
    # Order: more specific patterns first
    if "imperial ipa" in s or "double ipa" in s or "iipa" in s:
        return "Imperial IPA"
    if "hazy" in s and "ipa" in s or "neipa" in s or "new england ipa" in s:
        return "Hazy IPA"
    if "american ipa" in s or "ipa" in s and "american" in s:
        return "American IPA"
    if "english ipa" in s:
        return "English IPA"
    if "ipa" in s:
        return "American IPA"
    if "imperial stout" in s or "russian imperial" in s:
        return "Imperial Stout"
    if "milk stout" in s or "sweet stout" in s and "milk" in s:
        return "Milk Stout"
    if "oatmeal stout" in s:
        return "Oatmeal Stout"
    if "american stout" in s:
        return "American Stout"
    if "irish stout" in s or "dry stout" in s:
        return "Irish Stout"
    if "stout" in s:
        return "Stout"
    if "baltic porter" in s:
        return "Baltic Porter"
    if "robust porter" in s:
        return "Robust Porter"
    if "brown porter" in s:
        return "Brown Porter"
    if "porter" in s:
        return "Porter"
    if "gueuze" in s or "geuze" in s:
        return "Gueuze"
    if "lambic" in s:
        return "Lambic"
    if "flanders red" in s:
        return "Flanders Red"
    if "flanders" in s and "brown" in s:
        return "Flanders Brown"
    if "berliner weisse" in s or "berliner weiße" in s:
        return "Berliner Weisse"
    if "gose" in s:
        return "Gose"
    if "american wild" in s or "wild ale" in s:
        return "American Wild Ale"
    if "sour" in s or "saison" in s and "sour" in s:
        return "American Wild Ale"
    if "witbier" in s or "wit " in s or "belgian white" in s:
        return "Witbier"
    if "tripel" in s or "triple" in s and "belgian" in s:
        return "Belgian Tripel"
    if "dubbel" in s or "double" in s and "belgian" in s:
        return "Belgian Dubbel"
    if "belgian golden strong" in s or "belgian strong golden" in s:
        return "Belgian Golden Strong"
    if "belgian dark strong" in s or "quad" in s:
        return "Belgian Dark Strong"
    if "saison" in s:
        return "Saison"
    if "belgian blond" in s:
        return "Belgian Blond Ale"
    if "belgian pale" in s:
        return "Belgian Pale Ale"
    if "belgian" in s:
        return "Belgian Pale Ale"
    if "weissbier" in s or "hefeweizen" in s or "weizen" in s and "dunkel" not in s:
        return "Weissbier"
    if "dunkles weiss" in s or "dunkelweizen" in s:
        return "Dunkles Weissbier"
    if "weizenbock" in s:
        return "Weizenbock"
    if "american wheat" in s:
        return "American Wheat Beer"
    if "roggen" in s:
        return "Roggenbier"
    if "wheat" in s:
        return "American Wheat Beer"
    if "doppelbock" in s or "double bock" in s:
        return "Doppelbock"
    if "eisbock" in s:
        return "Eisbock"
    if "maibock" in s or "helles bock" in s:
        return "Maibock"
    if "dunkles bock" in s:
        return "Dunkles Bock"
    if "helles bock" in s:
        return "Helles Bock"
    if "bock" in s:
        return "Bock"
    if "cream ale" in s:
        return "Cream Ale"
    if "kolsch" in s or "kölsch" in s:
        return "Kolsch"
    if "rauchbier" in s or "smoked" in s:
        return "Rauchbier"
    if "marzen" in s or "märzen" in s or "oktoberfest" in s:
        return "Marzen"
    if "festbier" in s:
        return "Festbier"
    if "munich helles" in s or "helles" in s:
        return "Munich Helles"
    if "german pils" in s or "pilsner" in s and "german" in s:
        return "German Pils"
    if "czech premium" in s:
        return "Czech Premium Pale Lager"
    if "czech pale" in s:
        return "Czech Pale Lager"
    if "czech amber" in s:
        return "Czech Amber Lager"
    if "czech dark" in s:
        return "Czech Dark Lager"
    if "schwarzbier" in s:
        return "Schwarzbier"
    if "vienna lager" in s:
        return "Vienna Lager"
    if "international dark" in s:
        return "International Dark Lager"
    if "international amber" in s:
        return "International Amber Lager"
    if "international pale" in s:
        return "International Pale Lager"
    if "american lager" in s and "light" in s:
        return "American Light Lager"
    if "american lager" in s or "lager" in s and "american" in s:
        return "American Lager"
    if "pilsner" in s or "pils" in s:
        return "German Pils"
    if "lager" in s:
        return "American Lager"
    if "english barleywine" in s:
        return "English Barleywine"
    if "american barleywine" in s:
        return "American Barleywine"
    if "barleywine" in s or "barley wine" in s:
        return "American Barleywine"
    if "old ale" in s:
        return "Old Ale"
    if "winter warmer" in s:
        return "Winter Warmer"
    if "american strong" in s:
        return "American Strong Ale"
    if "strong ale" in s:
        return "American Strong Ale"
    if "irish red" in s:
        return "Irish Red Ale"
    if "scottish" in s:
        return "Scottish Export"
    if "altbier" in s or "alt" in s:
        return "Altbier"
    if "california common" in s or "steam beer" in s:
        return "California Common"
    if "american amber" in s:
        return "American Amber Ale"
    if "american brown" in s:
        return "American Brown Ale"
    if "english brown" in s:
        return "English Brown Ale"
    if "mild" in s:
        return "Dark Mild"
    if "brown ale" in s:
        return "American Brown Ale"
    if "best bitter" in s or "esb" in s:
        return "Strong Bitter"
    if "bitter" in s and "english" in s:
        return "Best Bitter"
    if "ordinary bitter" in s:
        return "Ordinary Bitter"
    if "english pale" in s:
        return "English Pale Ale"
    if "american pale" in s or "pale ale" in s:
        return "American Pale Ale"
    if "fruit" in s:
        return "Fruit Beer"
    if "spice" in s:
        return "Spice Beer"
    if "chocolate" in s and "stout" in s:
        return "Chocolate Stout"
    if "coffee" in s and "stout" in s:
        return "Coffee Stout"
    return None


def infer_style_from_name(name: str) -> str | None:
    """Infer style from beer name only (e.g. simple list)."""
    if not name:
        return None
    s = name.lower()
    if re.search(r'\bipa\b', s) and re.search(r'\bimperial\b|\bdouble\b', s):
        return "Imperial IPA"
    if re.search(r'\bipa\b', s):
        return "American IPA"
    if re.search(r'\bstout\b', s):
        return "Irish Stout"
    if re.search(r'\bporter\b', s):
        return "Porter"
    if re.search(r'\blager\b', s):
        return "American Lager"
    if re.search(r'\bpilsner\b|\bpils\b', s):
        return "German Pils"
    if re.search(r'\bwheat\b', s):
        return "American Wheat Beer"
    if re.search(r'\bbock\b', s):
        return "Bock"
    if re.search(r'\bpale ale\b', s):
        return "American Pale Ale"
    if re.search(r'\bamber\b', s):
        return "American Amber Ale"
    if re.search(r'\bbrown\b', s):
        return "American Brown Ale"
    return None


# -----------------------------------------------------------------------------
# 2C: Load fuzzy match mappings
# -----------------------------------------------------------------------------

def _data_path(*candidates: str) -> Path | None:
    """Return first path that exists under DATA_DIR."""
    for name in candidates:
        p = DATA_DIR / name
        if p.exists():
            return p
    return None


def load_beer_fuzzy_map() -> dict:
    mapping = {}
    path = _data_path("Beer_Name_Fuzzy_Match_List.csv", "Beer Name Fuzzy Match List.csv")
    if not path:
        print("  (Beer fuzzy match CSV not found, using empty mapping)")
        return mapping
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = normalize_name(row.get("Beer Name (Full)", ""))
            if key and "matches" in row:
                mapping[key] = row["matches"].strip()
    return mapping


def load_brewery_fuzzy_map() -> dict:
    mapping = {}
    path = _data_path("Brewery_Name_Fuzzy_Match_List.csv", "Brewery Name Fuzzy Match List.csv")
    if not path:
        print("  (Brewery fuzzy match CSV not found, using empty mapping)")
        return mapping
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = normalize_brewery(row.get("Brewery", ""))
            if key and "matches" in row:
                mapping[key] = row["matches"].strip()
    return mapping


# -----------------------------------------------------------------------------
# 2D: Load descriptor keywords
# -----------------------------------------------------------------------------

def load_descriptors() -> dict:
    path = _data_path("Beer_Descriptors_Simplified.xlsx", "Beer Descriptors Simplified.xlsx")
    descriptors = {"fruity": {}, "hoppy": {}, "spices": {}, "malty": {}}
    if not path:
        print("  (Beer_Descriptors_Simplified.xlsx not found, using empty descriptors)")
        return descriptors
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    cats = ["fruity", "hoppy", "spices", "malty"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        for i, cat in enumerate(cats):
            if i * 2 + 1 >= len(row):
                break
            keyword = row[i * 2]
            impact = row[i * 2 + 1]
            if keyword:
                descriptors[cat][str(keyword).strip().lower()] = int(impact or 1)
    wb.close()
    return descriptors


# -----------------------------------------------------------------------------
# 2E: PRIMARY — Ingest full_beer_reviews.xlsx
# -----------------------------------------------------------------------------

def ingest_full_reviews(brewery_fuzzy: dict) -> tuple[dict, dict]:
    """Read full_beer_reviews.xlsx, aggregate per beer_beerid. Returns (beers_dict, breweries_dict)."""
    path = DATA_DIR / "full_beer_reviews.xlsx"
    if not path.exists():
        print("ERROR: full_beer_reviews.xlsx not found in data/. Place the file and re-run.")
        sys.exit(1)
    print("  Reading full_beer_reviews.xlsx (expect 30-60 seconds)...")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    agg = defaultdict(lambda: {
        'name': '', 'brewery': '', 'brewery_id': None,
        'style': '', 'abv': None, 'beer_id': None, 'ratings': []
    })

    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        beer_id = row[6] if len(row) > 6 else None
        if not beer_id:
            continue
        b = agg[beer_id]
        b['beer_id'] = beer_id
        b['name'] = fix_mojibake(str(row[4] if len(row) > 4 else b['name'] or ''))
        b['brewery'] = fix_mojibake(str(row[1] if len(row) > 1 else b['brewery'] or ''))
        b['brewery_id'] = row[0] if len(row) > 0 else b['brewery_id']
        b['style'] = fix_mojibake(str(row[3] if len(row) > 3 else b['style'] or ''))
        if len(row) > 5 and row[5] is not None:
            b['abv'] = row[5]
        if len(row) > 2 and row[2] is not None:
            b['ratings'].append(float(row[2]))

    wb.close()
    print(f"  Processed {row_count} review rows -> {len(agg)} unique beers")
    if row_count >= 1048575:
        print("  WARNING: Hit Excel row limit (1,048,575). Dataset may be truncated.")

    breweries = {}
    beers = {}

    for beer_id, b in agg.items():
        beer_name = (b['name'] or '').strip()
        brewery_raw = (b['brewery'] or '').strip()
        if not beer_name:
            continue
        norm_brew = normalize_brewery(brewery_raw)
        canonical_brewery = brewery_fuzzy.get(norm_brew, brewery_raw)
        brew_key = normalize_brewery(canonical_brewery)

        if brew_key and brew_key not in breweries:
            breweries[brew_key] = {
                "name": canonical_brewery,
                "slug": slugify(canonical_brewery),
                "normalized_name": brew_key,
                "source": "full_reviews",
                "source_id": str(b['brewery_id']) if b['brewery_id'] else None,
            }

        ratings = b['ratings']
        avg_overall = round(sum(ratings) / len(ratings), 2) if ratings else None
        style_raw = fix_mojibake((b['style'] or '').strip()) or None
        beer_dedup = f"{brew_key} {normalize_name(beer_name)}"

        beers[beer_dedup] = {
            "name": beer_name,
            "slug": slugify(f"{canonical_brewery} {beer_name}"),
            "normalized_name": normalize_name(beer_name),
            "brewery_key": brew_key,
            "brewery_name": canonical_brewery,
            "style": style_raw,
            "style_category": map_style_to_bjcp(style_raw),
            "abv": safe_float(b['abv']),
            "description": None,
            "ibu_min": None,
            "ibu_max": None,
            "flavor_astringency": None,
            "flavor_body": None,
            "flavor_alcohol": None,
            "flavor_bitter": None,
            "flavor_sweet": None,
            "flavor_sour": None,
            "flavor_salty": None,
            "flavor_fruity": None,
            "flavor_hoppy": None,
            "flavor_spicy": None,
            "flavor_malty": None,
            "review_aroma": None,
            "review_appearance": None,
            "review_palate": None,
            "review_taste": None,
            "review_overall": avg_overall,
            "review_count": len(ratings),
            "source": "full_reviews",
            "source_id": str(beer_id),
            "source_brewery_id": str(b['brewery_id']) if b['brewery_id'] else None,
        }

    return beers, breweries


# -----------------------------------------------------------------------------
# 2F: ENRICHMENT — Merge beer_profile_and_ratings.csv
# -----------------------------------------------------------------------------

def merge_profiles(beers: dict, breweries: dict, beer_fuzzy: dict, brewery_fuzzy: dict) -> None:
    path = _data_path("beer_profile_and_ratings.csv")
    if not path:
        print("  (beer_profile_and_ratings.csv not found, skipping profiles)")
        return
    enriched = 0
    added = 0
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            beer_name = (row.get("Name") or "").strip()
            brewery_raw = (row.get("Brewery") or "").strip()
            style = (row.get("Style") or "").strip()
            canonical_beer = beer_fuzzy.get(normalize_name(beer_name), beer_name)
            norm_brew = normalize_brewery(brewery_raw)
            canonical_brewery = brewery_fuzzy.get(norm_brew, brewery_raw)
            brew_key = normalize_brewery(canonical_brewery)
            beer_dedup = f"{brew_key} {normalize_name(canonical_beer)}"

            if brew_key and brew_key not in breweries:
                breweries[brew_key] = {
                    "name": canonical_brewery,
                    "slug": slugify(canonical_brewery),
                    "normalized_name": brew_key,
                    "source": "profile",
                    "source_id": None,
                }

            description_raw = (row.get("Description") or "").strip()
            description = re.sub(r'^Notes:\s*', '', description_raw).rstrip('\t').strip() or None

            enrichment = {
                "description": description,
                "ibu_min": safe_int(row.get("Min IBU")),
                "ibu_max": safe_int(row.get("Max IBU")),
                "flavor_astringency": safe_int(row.get("Astringency")),
                "flavor_body": safe_int(row.get("Body")),
                "flavor_alcohol": safe_int(row.get("Alcohol")),
                "flavor_bitter": safe_int(row.get("Bitter")),
                "flavor_sweet": safe_int(row.get("Sweet")),
                "flavor_sour": safe_int(row.get("Sour")),
                "flavor_salty": safe_int(row.get("Salty")),
                "flavor_fruity": safe_int(row.get("Fruits")),
                "flavor_hoppy": safe_int(row.get("Hoppy")),
                "flavor_spicy": safe_int(row.get("Spices")),
                "flavor_malty": safe_int(row.get("Malty")),
                "review_aroma": safe_float(row.get("review_aroma")),
                "review_appearance": safe_float(row.get("review_appearance")),
                "review_palate": safe_float(row.get("review_palate")),
                "review_taste": safe_float(row.get("review_taste")),
            }

            if beer_dedup in beers:
                existing = beers[beer_dedup]
                for key, val in enrichment.items():
                    if val is not None and existing.get(key) is None:
                        existing[key] = val
                if style and not existing.get("style"):
                    existing["style"] = style
                    existing["style_category"] = map_style_to_bjcp(style)
                if safe_float(row.get("ABV")) and not existing.get("abv"):
                    existing["abv"] = safe_float(row.get("ABV"))
                enriched += 1
            else:
                beers[beer_dedup] = {
                    "name": beer_name,
                    "slug": slugify(f"{canonical_brewery} {beer_name}"),
                    "normalized_name": normalize_name(beer_name),
                    "brewery_key": brew_key,
                    "brewery_name": canonical_brewery,
                    "style": style or None,
                    "style_category": map_style_to_bjcp(style),
                    "abv": safe_float(row.get("ABV")),
                    "source": "profile",
                    "source_id": None,
                    "source_brewery_id": None,
                    "review_overall": safe_float(row.get("review_overall")),
                    "review_count": safe_int(row.get("number_of_reviews")),
                    **enrichment,
                }
                added += 1
    print(f"  Profiles: enriched {enriched} existing beers, added {added} new")


# -----------------------------------------------------------------------------
# 2J: Helper functions
# -----------------------------------------------------------------------------

def make_empty_beer(beer_name, brew_key, brewery_name, source, style=None, full_name=None):
    slug = slugify(full_name if full_name else (f"{brewery_name} {beer_name}" if brewery_name else beer_name))
    return {
        "name": beer_name,
        "slug": slug,
        "normalized_name": normalize_name(beer_name),
        "brewery_key": brew_key,
        "brewery_name": brewery_name,
        "style": style,
        "style_category": map_style_to_bjcp(style) if style else None,
        "abv": None,
        "description": None,
        "ibu_min": None,
        "ibu_max": None,
        "flavor_astringency": None,
        "flavor_body": None,
        "flavor_alcohol": None,
        "flavor_bitter": None,
        "flavor_sweet": None,
        "flavor_sour": None,
        "flavor_salty": None,
        "flavor_fruity": None,
        "flavor_hoppy": None,
        "flavor_spicy": None,
        "flavor_malty": None,
        "review_aroma": None,
        "review_appearance": None,
        "review_palate": None,
        "review_taste": None,
        "review_overall": None,
        "review_count": None,
        "source": source,
        "source_id": None,
        "source_brewery_id": None,
    }


def try_extract_brewery(full_name: str, known_breweries: list) -> tuple:
    lower = full_name.lower()
    for brewery in known_breweries:
        if lower.startswith(brewery.lower()) and len(brewery) > 3:
            remainder = full_name[len(brewery):].strip()
            if remainder:
                return brewery, remainder
    return None, full_name


# -----------------------------------------------------------------------------
# 2G: Tertiary — manufacturers + simple list
# -----------------------------------------------------------------------------

def ingest_manufacturers(beers: dict, breweries: dict, brewery_fuzzy: dict) -> None:
    path = _data_path("beermanufacturersmicrobrewersbrands.csv")
    if not path:
        print("  (beermanufacturersmicrobrewersbrands.csv not found, skipping)")
        return
    added = 0
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            brewery_raw = (row.get("Brewery") or "").strip()
            beer_name = (row.get("Beer Name") or "").strip()
            if not beer_name:
                continue
            norm_brew = normalize_brewery(brewery_raw)
            canonical_brewery = brewery_fuzzy.get(norm_brew, brewery_raw)
            brew_key = normalize_brewery(canonical_brewery)
            beer_dedup = f"{brew_key} {normalize_name(beer_name)}"

            if brew_key and brew_key not in breweries:
                breweries[brew_key] = {
                    "name": canonical_brewery,
                    "slug": slugify(canonical_brewery),
                    "normalized_name": brew_key,
                    "source": "manufacturer",
                    "source_id": None,
                }
            if beer_dedup not in beers:
                beers[beer_dedup] = make_empty_beer(beer_name, brew_key, canonical_brewery, "manufacturer")
                added += 1
    print(f"  Manufacturers: +{added} new beers")


def ingest_simple_list(beers: dict, breweries: dict, brewery_fuzzy: dict) -> None:
    path = _data_path("beer_list_simple.txt", "beer list simple.txt")
    if not path:
        print("  (beer_list_simple.txt not found, skipping)")
        return
    added = 0
    known = sorted(set(b["name"] for b in breweries.values()), key=len, reverse=True)
    with open(path, encoding="utf-8") as f:
        for line in f:
            name = line.strip()
            if not name:
                continue
            dedup_key = normalize_name(name)
            if dedup_key in beers:
                continue
            brewery, beer = try_extract_brewery(name, known)
            brew_key = normalize_brewery(brewery) if brewery else ""
            beers[dedup_key] = make_empty_beer(
                beer or name,
                brew_key,
                brewery or "",
                "simple_list",
                style=infer_style_from_name(name),
                full_name=name,
            )
            added += 1
    print(f"  Simple list: +{added} new beers")


# -----------------------------------------------------------------------------
# 2K: Output writers
# -----------------------------------------------------------------------------

def write_breweries_csv(breweries: dict) -> None:
    path = OUT_DIR / "breweries.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "slug", "normalized_name", "source", "source_id"])
        for r in breweries.values():
            w.writerow([
                csv_val(r["name"]),
                csv_val(r["slug"]),
                csv_val(r["normalized_name"]),
                csv_val(r["source"]),
                csv_val(r.get("source_id")),
            ])
    print(f"  Wrote {path}")


def write_beers_csv(beers: dict) -> None:
    path = OUT_DIR / "beers.csv"
    cols = [
        "name", "slug", "normalized_name", "brewery_normalized_name", "brewery_name",
        "style", "style_category", "abv", "ibu_min", "ibu_max",
        "flavor_astringency", "flavor_body", "flavor_alcohol", "flavor_bitter",
        "flavor_sweet", "flavor_sour", "flavor_salty", "flavor_fruity",
        "flavor_hoppy", "flavor_spicy", "flavor_malty",
        "review_aroma", "review_appearance", "review_palate", "review_taste",
        "review_overall", "review_count",
        "description", "source", "source_id", "source_brewery_id",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for r in beers.values():
            w.writerow([
                csv_val(r.get("name")),
                csv_val(r.get("slug")),
                csv_val(r.get("normalized_name")),
                csv_val(r.get("brewery_key")),
                csv_val(r.get("brewery_name")),
                csv_val(r.get("style")),
                csv_val(r.get("style_category")),
                csv_val(r.get("abv")),
                csv_val(r.get("ibu_min")),
                csv_val(r.get("ibu_max")),
                csv_val(r.get("flavor_astringency")),
                csv_val(r.get("flavor_body")),
                csv_val(r.get("flavor_alcohol")),
                csv_val(r.get("flavor_bitter")),
                csv_val(r.get("flavor_sweet")),
                csv_val(r.get("flavor_sour")),
                csv_val(r.get("flavor_salty")),
                csv_val(r.get("flavor_fruity")),
                csv_val(r.get("flavor_hoppy")),
                csv_val(r.get("flavor_spicy")),
                csv_val(r.get("flavor_malty")),
                csv_val(r.get("review_aroma")),
                csv_val(r.get("review_appearance")),
                csv_val(r.get("review_palate")),
                csv_val(r.get("review_taste")),
                csv_val(r.get("review_overall")),
                csv_val(r.get("review_count")),
                csv_val(r.get("description")),
                csv_val(r.get("source")),
                csv_val(r.get("source_id")),
                csv_val(r.get("source_brewery_id")),
            ])
    print(f"  Wrote {path}")


def write_beer_styles_csv() -> None:
    path = OUT_DIR / "beer_styles.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "category", "description", "abv_min", "abv_max", "ibu_min", "ibu_max"])
        for row in BJCP_STYLES:
            name, category, description, abv_min, abv_max, ibu_min, ibu_max = row
            w.writerow([
                csv_val(name),
                csv_val(category),
                csv_val(description),
                csv_val(abv_min),
                csv_val(abv_max),
                csv_val(ibu_min),
                csv_val(ibu_max),
            ])
    print(f"  Wrote {path}")


def write_flavor_descriptors_csv(descriptors: dict) -> None:
    path = OUT_DIR / "flavor_descriptors.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["category", "keyword", "impact"])
        for cat, kws in descriptors.items():
            for keyword, impact in kws.items():
                w.writerow([csv_val(cat), csv_val(keyword), csv_val(impact)])
    print(f"  Wrote {path}")


# -----------------------------------------------------------------------------
# 2L: Main pipeline + stats
# -----------------------------------------------------------------------------

def main() -> None:
    print("=" * 50)
    print("Beer Catalog Pipeline")
    print("=" * 50)

    brewery_fuzzy = load_brewery_fuzzy_map()
    beer_fuzzy = load_beer_fuzzy_map()
    descriptors = load_descriptors()

    beers, breweries = ingest_full_reviews(brewery_fuzzy)
    merge_profiles(beers, breweries, beer_fuzzy, brewery_fuzzy)
    ingest_manufacturers(beers, breweries, brewery_fuzzy)
    ingest_simple_list(beers, breweries, brewery_fuzzy)

    write_breweries_csv(breweries)
    write_beers_csv(beers)
    write_beer_styles_csv()
    write_flavor_descriptors_csv(descriptors)

    print(f"\n{'=' * 50}")
    print("FINAL STATS")
    print(f"{'=' * 50}")
    print(f"Breweries:         {len(breweries)}")
    print(f"Beers:             {len(beers)}")
    print(f"  with ABV:        {sum(1 for b in beers.values() if b.get('abv'))}")
    print(f"  with style:      {sum(1 for b in beers.values() if b.get('style'))}")
    print(f"  with reviews:    {sum(1 for b in beers.values() if b.get('review_count') and b['review_count'] > 0)}")
    print(f"  with flavors:    {sum(1 for b in beers.values() if b.get('flavor_hoppy') is not None)}")
    print(f"  with description:{sum(1 for b in beers.values() if b.get('description'))}")
    print(f"Styles:            {len(BJCP_STYLES)}")
    print(f"Descriptors:       {sum(len(v) for v in descriptors.values())}")
    print("\nBy source:")
    for src, cnt in Counter(b["source"] for b in beers.values()).most_common():
        print(f"  {src}: {cnt}")


if __name__ == "__main__":
    main()
